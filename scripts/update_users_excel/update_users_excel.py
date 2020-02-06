#!/usr/bin/env python
"""
Copyright 2019 David Wong

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

from collections import OrderedDict
from enum import Enum
from copy import deepcopy
# from pathlib import Path
from contextlib import closing
from warnings import warn
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO, StringIO
import argparse
import abc
import csv
import datetime
import mysql.connector

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.writer import theme, excel

import requests
from urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)


class ConfigType(Enum):
    DATABASE = "database"
    WIKI = "wiki"

    def __str__(self):
        return self.value


USERS_EXCEL_FILE_NAME = "Users.xlsx"
USERS_EXCEL_FILE_PATH = "./" + USERS_EXCEL_FILE_NAME

USER_FIELD_TITLE = OrderedDict([
    ("user_name", "Username"),
    ("user_real_name", "Real name"),
    ("user_email", "Email"),
    ("user_registration", "Registration date")
])

CONFIG = {
    "database_host": "192.168.56.56",
    "database_port": 3306,
    "database_database": "wiki_demo",
    "database_username": "wiki_app_user",
    "database_password": "password",
    "database_table_prefix": "",
    "database_user_table": "user",
    "database_user_field_title": USER_FIELD_TITLE,

    "wiki_uri": "https://192.168.56.56/demo",
    "wiki_api_path": "/api.php",
    "wiki_username": "Admin",
    "wiki_password": "adminpass",
    "wiki_user_field_title": USER_FIELD_TITLE,

    "config_type": ConfigType.WIKI,
    "users_excel_file_name": USERS_EXCEL_FILE_NAME,
    "users_excel_file_path": USERS_EXCEL_FILE_PATH
}


class UserModel:
    def __init__(
        self,
        field_title=None,
        fields=None,
        titles=None
    ):
        self.field_title, self.fields, self.titles = self.parse_fields_and_titles(field_title, fields, titles)

    @staticmethod
    def parse_fields_and_titles(field_title=None, fields=None, titles=None):
        is_field_title_dict = isinstance(field_title, dict)
        is_fields_list = isinstance(fields, list)
        is_titles_list = isinstance(titles, list)

        if is_field_title_dict:
            if is_fields_list:
                warn("`field_title` dictionary already supplied. Ignoring `fields` list...")
            fields = list(field_title.keys())

            if is_titles_list:
                warn("`field_title` dictionary already supplied. Ignoring `titles` list...")
            titles = list(field_title.values())
        else:
            if not is_fields_list:
                raise TypeError("Must supply either `field_title` dictionary or `fields` list.")

            if titles is None:
                titles = fields
            elif not is_titles_list:
                raise TypeError("`titles` must be a list.")

            number_of_fields = len(fields)
            number_of_titles = len(titles)
            if number_of_fields > number_of_titles:
                warn("There are more fields than titles.")
            elif number_of_titles > number_of_fields:
                warn("There are more titles than fields.")

            field_title = OrderedDict([
                (fields[i], (titles if i < number_of_titles else fields)[i])
                for i in range(number_of_fields)
            ])

        return field_title, fields, titles

    @staticmethod
    def format_date(timestamp):
        """
        https://www.mediawiki.org/wiki/Manual:Timestamp
        :param timestamp:
        :return:
        """

        date = datetime.datetime.strptime(timestamp, "%Y%m%d%H%M%S")
        # return date.isoformat()
        return date

    @classmethod
    def format_user_date(cls, user, key):
        user[key] = cls.format_date(user[key])
        return user

    @classmethod
    def format_user_registration_date(cls, user):
        return cls.format_user_date(user, "user_registration")

    @classmethod
    def format_user_dates(cls, users):
        for user in users:
            cls.format_user_registration_date(user)
        return users


class UserController(metaclass=abc.ABCMeta):
    def __init__(self, user_model):
        if not isinstance(user_model, UserModel):
            raise TypeError("`user_model` must be a UserModel instance.")
        self.user_model = user_model

    @abc.abstractmethod
    def fetch_users(self):
        raise NotImplementedError("`fetch_users` must be implemented.")

    def fetch_formatted_users(self):
        users = self.fetch_users()
        self.user_model.format_user_dates(users)
        return users


class DatabaseUserModel(UserModel):
    def __init__(
        self,
        database_model,

        table="user",
        field_title=None,
        fields=None,
        titles=None
    ):
        super().__init__(field_title, fields, titles)

        if not isinstance(database_model, DatabaseModel):
            raise TypeError("`database_model` must be a DatabaseModel instance.")
        self.database_model = database_model

        self.table = database_model.config_model.table_prefix + table

        self.query = f"""
            select {", ".join(self.fields)}
            from `{self.table}`;
        """


class ExportUserModel(UserModel):
    def __init__(
        self,
        field_title=None,
        fields=None,
        titles=None
    ):
        super().__init__(field_title, fields, titles)


class DatabaseUserController(UserController):
    def __init__(self, database_model, database_controller):
        if not isinstance(database_model, DatabaseModel):
            raise TypeError("`database_model` must be a DatabaseModel instance.")
        self.database_model = database_model

        if not isinstance(database_controller, DatabaseController):
            raise TypeError("`database_controller` must be a DatabaseController instance.")
        self.database_controller = database_controller

        super().__init__(database_model.user_model)

    def fetch_users(self):
        return self.database_controller.execute(self.database_model.user_model.query)


class ExportUserController(UserController):
    def __init__(self, export_model, export_controller):
        if not isinstance(export_model, ExportModel):
            raise TypeError("`export_model` must be a ExportModel instance.")
        self.export_model = export_model

        if not isinstance(export_controller, ExportController):
            raise TypeError("`export_controller` must be a ExportController instance.")
        self.export_controller = export_controller

        super().__init__(export_model.user_model)

    def fetch_users_csv(self, token=None):
        export_model = self.export_model

        if not isinstance(token, str):
            token = self.export_controller.wiki_controller.csrf_token

        body = {
            "title": export_model.user_export_title,
            "exportusers": 1,
            "token": token,
            "wpEditToken": token
        }

        response = self.export_controller.wiki_controller.session.post(url=export_model.user_export_uri, data=body)

        data = response.content

        return data

    def fetch_users(self):
        data = self.fetch_users_csv()
        text = data.decode("utf-8")
        reader = csv.DictReader(StringIO(text), delimiter=",", quotechar="\"")
        return list(reader)


class DatabaseConfigModel:
    def __init__(
        self,
        username,
        password,
        host,
        port=3306,
        database=None,
        table_prefix=""
    ):
        self.config_type = ConfigType.DATABASE

        self.username = username
        self.password = password
        self.host = host
        self.port = port
        self.database = database

        self.table_prefix = table_prefix


class DatabaseModel:
    def __init__(
        self,

        config_model=None,
        username=None,
        password=None,
        host=None,
        port=3306,
        database=None,
        table_prefix="",

        user_model=None,
        user_table="user",
        user_field_title=None,
        user_fields=None,
        user_titles=None
    ):
        self.config_type = ConfigType.DATABASE

        self.config_model = (config_model if isinstance(config_model, DatabaseConfigModel) else DatabaseConfigModel(
            username=username,
            password=password,
            host=host,
            port=port,
            database=database,
            table_prefix=table_prefix
        ))

        self.user_model = (user_model if isinstance(user_model, DatabaseUserModel) else DatabaseUserModel(
            self, table=user_table, field_title=user_field_title, fields=user_fields, titles=user_titles
        ))


class WikiConfigModel:
    def __init__(
        self,
        uri,
        api_path,
        username,
        password
    ):
        self.uri = uri
        self.api_path = api_path
        self.api_endpoint = uri + api_path

        self.username = username
        self.password = password


class WikiController:
    def __init__(self, wiki_config_model):
        if not isinstance(wiki_config_model, WikiConfigModel):
            raise TypeError("`wiki_config_model` must be a WikiConfigModel instance.")
        self.config_model = wiki_config_model

        session = requests.Session()
        session.verify = False
        self.session = session

        self.current_login_token = None
        self.current_csrf_token = None

    def fetch_tokens(self, token_type):
        body = {
            "action": "query",
            "meta": "tokens",
            "type": token_type,
            "format": "json"
        }

        response = self.session.get(url=self.config_model.api_endpoint, params=body)
        data = response.json()

        tokens = data["query"]["tokens"]

        return tokens

    def fetch_login_token(self):
        return self.fetch_tokens("login")["logintoken"]

    def fetch_csrf_token(self):
        return self.fetch_tokens("csrf")["csrftoken"]

    @property
    def login_token(self):
        token = self.current_login_token
        if not isinstance(token, str):
            token = self.fetch_login_token()
            self.current_login_token = token
        return token

    @property
    def csrf_token(self):
        token = self.current_csrf_token
        if not isinstance(token, str):
            token = self.fetch_csrf_token()
            self.current_csrf_token = token
        return token

    def login(self, token=None, return_uri=None):
        config_model = self.config_model
        username = config_model.username
        password = config_model.password

        if not isinstance(token, str):
            token = self.login_token

        if not isinstance(return_uri, str):
            return_uri = config_model.uri

        body = {
            "action": "clientlogin",
            "username": username,
            "password": password,
            "loginreturnurl": return_uri,
            "logintoken": token,
            "format": "json"
        }

        response = self.session.post(url=config_model.api_endpoint, data=body)

        data = response.json()

        return data

    def upload_file(self, file_name, file_data, token=None):
        if not isinstance(token, str):
            token = self.csrf_token

        body = {
            "action": "upload",
            "filename": file_name,
            "token": token,
            "format": "json",
            "ignorewarnings": 1
        }

        files = {
            "file": (file_name, file_data, "multipart/form-data")
        }

        response = self.session.post(url=self.config_model.api_endpoint, files=files, data=body)

        data = None

        try:
            data = response.json()

            print(data)
        except ValueError:
            print(response)
            print(response.content)

        return data


class ExportModel:
    def __init__(
        self,

        wiki_model=None,
        uri=None,
        api_path=None,
        username=None,
        password=None,

        user_model=None,
        user_field_title=None,
        user_fields=None,
        user_titles=None
    ):
        self.config_type = ConfigType.WIKI

        self.wiki_model = (wiki_model if isinstance(wiki_model, WikiConfigModel) else WikiConfigModel(
            uri=uri,
            api_path=api_path,
            username=username,
            password=password
        ))

        user_export_title = "Special:Userexport"
        self.user_export_title = user_export_title
        self.user_export_uri = self.wiki_model.uri + "/index.php/" + user_export_title

        self.user_model = (user_model if isinstance(user_model, ExportUserModel) else ExportUserModel(
            field_title=user_field_title, fields=user_fields, titles=user_titles
        ))


class DatabaseController:
    def __init__(self, database_model, wiki_controller=None):
        self.config_type = ConfigType.DATABASE

        if not isinstance(database_model, DatabaseModel):
            raise TypeError("`database_model` must be a DatabaseModel instance.")
        self.database_model = database_model

        self.database_user_controller = DatabaseUserController(database_model, self)

    def execute(self, query):
        config_model = self.database_model.config_model
        rows = []
        with closing(mysql.connector.connect(
            user=config_model.username,
            password=config_model.password,
            host=config_model.host,
            port=config_model.port,
            database=config_model.database
        )) as connection:
            with closing(connection.cursor(dictionary=True)) as cursor:
                cursor.execute(query)
                result = cursor.fetchall()
                rows = list(result)
        return rows

    def fetch_users(self):
        return self.database_user_controller.fetch_users()

    def fetch_formatted_users(self):
        return self.database_user_controller.fetch_formatted_users()


class ExportController:
    def __init__(self, export_model, wiki_controller=None):
        self.config_type = ConfigType.WIKI

        if not isinstance(export_model, ExportModel):
            raise TypeError("`export_model` must be a ExportModel instance.")
        self.export_model = export_model

        self.export_user_controller = ExportUserController(export_model, self)

        self.wiki_controller = (wiki_controller if isinstance(wiki_controller, WikiController) else WikiController(export_model.wiki_model))
        wiki_controller.login()

    def fetch_users(self):
        return self.export_user_controller.fetch_users()

    def fetch_formatted_users(self):
        return self.export_user_controller.fetch_formatted_users()


class WorkbookController:
    XL_FOLDER_NAME = "xl"

    CONTENT_TYPES_XML_FILE_NAME = "[Content_Types].xml"
    WORKBOOK_XML_FILE_NAME = "workbook.xml"
    STYLES_XML_FILE_NAME = "styles.xml"

    FIRST_NAMES = [
        CONTENT_TYPES_XML_FILE_NAME,
        f"{XL_FOLDER_NAME}/{WORKBOOK_XML_FILE_NAME}",
        f"{XL_FOLDER_NAME}/{STYLES_XML_FILE_NAME}"
    ]

    ISO_8601_NUMBER_FORMAT = "yyyy-mm-ddThh:MM:ss"

    was_theme_updated = False

    def __init__(self):
        cls = self.__class__
        if not cls.was_theme_updated:
            self.update_theme()
            cls.was_theme_updated = True

    @staticmethod
    def update_theme():
        color = {
            "1F497D": "44546A",
            "EEECE1": "E7E6E6",
            "4F81BD": "5B9BD5",
            "C0504D": "ED7D31",
            "9BBB59": "A5A5A5",
            "8064A2": "FFC000",
            "4BACC6": "4472C4",
            "F79646": "70AD47",
            "0000FF": "0563C1",
            "800080": "954F72"
        }

        xml = theme.theme_xml
        for original, replacement in color.items():
            xml = xml.replace(f"val=\"{original}\"", f"val=\"{replacement}\"")

        theme.theme_xml = xml
        excel.theme_xml = xml

    @classmethod
    def fix_workbook_mime_type(cls, file_path):
        buffer = BytesIO()

        with ZipFile(file_path) as zip_file:
            names = zip_file.namelist()

            FIRST_NAMES = cls.FIRST_NAMES
            remaining_names = [name for name in names if name not in FIRST_NAMES]
            ordered_names = FIRST_NAMES + remaining_names

            with ZipFile(buffer, "w", ZIP_DEFLATED, allowZip64=True) as buffer_zip_file:
                for name in ordered_names:
                    try:
                        file = zip_file.open(name)
                        buffer_zip_file.writestr(file.name, file.read())
                    except KeyError:
                        pass

        return buffer

    @staticmethod
    def autosize_columns(worksheet):
        def value_of(value):
            return (str(value) if value is not None else "")

        for cells in worksheet.columns:
            length = max(len(value_of(cell.value)) for cell in cells)
            column_letter = get_column_letter(cells[0].column)
            worksheet.column_dimensions[column_letter].width = length

        return worksheet

    @classmethod
    def create_users_workbook(cls, field_title, fields, titles, users):
        # Initialize workbook and worksheet.
        workbook = Workbook()
        properties = workbook.properties
        properties.title = "Users"
        properties.creator = None

        sheet = workbook.active
        sheet.title = "Users"

        table_name = "User"

        # Add data.
        rows = [field_title] + users
        for r, row in enumerate(rows, start=1):
            for c, field in enumerate(fields, start=1):
                value = row[field]
                cell = sheet.cell(row=r, column=c)

                cell_number_format = FORMAT_TEXT
                cell_value = value
                cell_data_type = "s"
                if isinstance(value, datetime.datetime):
                    cell_number_format = cls.ISO_8601_NUMBER_FORMAT
                    cell_data_type = "d"

                cell.number_format = cell_number_format
                cell.value = cell_value
                cell.data_type = cell_data_type

        total_row = {
            "user_name": "Total",
            "user_real_name": "",
            "user_email": "",
            "user_registration": f"=SUBTOTAL(103,{table_name}[{titles[-1]}])"
        }
        sheet.append([total_row[field] for field in fields])

        # Add table.
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table_columns = tuple(TableColumn(id=h, name=header) for h, header in enumerate(titles, start=1))
        total_column = table_columns[0]
        total_column.totalsRowLabel = "Total"
        count_column = table_columns[-1]
        count_column.totalsRowFunction = "count"

        max_column = sheet.max_column
        max_column_letter = get_column_letter(max_column)
        max_row = sheet.max_row
        table = Table(
            displayName=table_name,
            ref=f"A1:{max_column_letter}{max_row}",
            autoFilter=AutoFilter(ref=f"A1:{max_column_letter}{max_row - 1}"),
            tableStyleInfo=style,
            totalsRowShown=True,
            totalsRowCount=1,
            tableColumns=table_columns
        )

        sheet.add_table(table)

        # Adjust column sizes.
        cls.autosize_columns(sheet)

        # Set the active cell under the table.
        active_cell = f"A{max_row + 1}"
        selection = sheet.sheet_view.selection[0]
        selection.activeCell = active_cell
        selection.sqref = active_cell

        return workbook


class ArgumentController:
    def __init__(self, default_config):
        self.default_config = default_config
        self.parser = None
        self.config = None

    @staticmethod
    def add_argument(parser, *args, **kwargs):
        if "metavar" not in kwargs and "choices" not in kwargs and "default" in kwargs:
            kwargs["metavar"] = kwargs["default"]
        parser.add_argument(*args, **kwargs)

    @classmethod
    def add_arguments(cls, parser, arguments):
        add_argument = cls.add_argument
        for argument in arguments:
            names, kwargs = argument
            add_argument(parser, *names, **kwargs)

    def create_parser(self):
        parser = argparse.ArgumentParser(description="Fetches the list of users from a database or wiki, creates an Excel workbook, and then uploads the Excel file onto the wiki.")

        config = self.default_config
        add_arguments = self.add_arguments

        database_group = parser.add_argument_group("database", "Database config.")
        add_arguments(database_group, [
            (["--database-host", "--db-host"], {"help": "Database host.", "default": config["database_host"]}),
            (["--database-port", "--db-port"], {"help": "Database port.", "type": int, "default": config["database_port"]}),
            (["--database-database", "--db-database", "--db-db"], {"help": "Database database.", "default": config["database_database"]}),
            (["--database-username", "--db-username", "--db-user"], {"help": "Database username.", "default": config["database_username"]}),
            (["--database-password", "--db-password", "--db-pass"], {"help": "Database password.", "default": config["database_password"]}),
            (["--database-table-prefix", "--db-table-prefix"], {"help": "Database table prefix.", "default": config["database_table_prefix"]}),
            (["--database-user-table", "--db-user-table"], {"help": "Database user table.", "default": config["database_user_table"]})
        ])

        wiki_group = parser.add_argument_group("wiki", "Wiki config.")
        add_arguments(wiki_group, [
            (["--wiki-uri", "--w-uri"], {"help": "Wiki URI.", "default": config["wiki_uri"]}),
            (["--wiki-api-path", "--w-api-path"], {"help": "Wiki API path.", "default": config["wiki_api_path"]}),
            (["--wiki-username", "--w-username", "--w-user"], {"help": "Wiki username.", "default": config["wiki_username"]}),
            (["--wiki-password", "--w-password", "--w-pass"], {"help": "Wiki password.", "default": config["wiki_password"]})
        ])

        add_arguments(parser, [
            (["--config-type"], {"help": "Config type.", "type": ConfigType, "choices": list(ConfigType), "default": config["config_type"]}),
            (["--users-excel-file-name"], {"help": "Users Excel file name.", "default": config["users_excel_file_name"]}),
            (["--users-excel-file-path"], {"help": "Users Excel file path.", "default": config["users_excel_file_path"]})
        ])

        self.parser = parser

        return parser

    def parse_system_arguments(self):
        parser = self.create_parser()
        result = vars(parser.parse_args())
        return result

    def create_config_from_arguments(self, result):
        config = deepcopy(self.default_config)

        for key, value in result.items():
            config[key] = value

        self.config = config

        return config

    def create_config_from_system_arguments(self):
        result = self.parse_system_arguments()
        return self.create_config_from_arguments(result)


class ConfigModel:
    def __init__(self, config):
        self.database_config = {
            "host": config["database_host"],
            "port": config["database_port"],
            "database": config["database_database"],
            "username": config["database_username"],
            "password": config["database_password"],
            "table_prefix": config["database_table_prefix"],
            "user_table": config["database_user_table"],
            "user_field_title": config["database_user_field_title"],
        }

        self.wiki_config = {
            "uri": config["wiki_uri"],
            "api_path": config["wiki_api_path"],
            "username": config["wiki_username"],
            "password": config["wiki_password"],
            "user_field_title": config["wiki_user_field_title"]
        }

        self.config_type = config["config_type"]
        self.users_excel_file_name = config["users_excel_file_name"]
        self.users_excel_file_path = config["users_excel_file_path"]


class MainController:
    def __init__(self, config_model):
        self.config_model = (config_model if isinstance(config_model, ConfigModel) else ConfigModel(config_model))

        self.user_model = None
        self.users = None
        self.wiki_controller = None
        self.workbook_buffer = None

    def run(self):
        config_type = self.config_model.config_type
        print(f"Fetching users from {config_type.value}...", end="")
        users = (self.fetch_users_from_export() if config_type == ConfigType.WIKI else self.fetch_users_from_database())
        self.users = users
        print(" Dome.")
        # print(users)
        print()

        print("Creating users workbook...", end="")
        self.create_users_workbook()
        print(" Dome.")
        print("Writing users workbook...", end="")
        self.write_users_workbook()
        print(" Dome.")

        print("Uploading users workbook...")
        self.upload_users_workbook()

    def fetch_users(self, source_config, Model, Controller):
        model = Model(**source_config)

        self.user_model = model.user_model

        wiki_controller = self.wiki_controller
        if not isinstance(wiki_controller, WikiController):
            wiki_config = self.config_model.wiki_config
            wiki_model = (model.wiki_model if hasattr(model, "wiki_model") else WikiConfigModel(
                uri=wiki_config["uri"],
                api_path=wiki_config["api_path"],
                username=wiki_config["username"],
                password=wiki_config["password"]
            ))
            wiki_controller = WikiController(wiki_model)
            wiki_controller.login()
            self.wiki_controller = wiki_controller

        controller = Controller(model, wiki_controller)
        users = controller.fetch_formatted_users()
        return users

    def fetch_users_from_database(self):
        return self.fetch_users(self.config_model.database_config, DatabaseModel, DatabaseController)

    def fetch_users_from_export(self):
        return self.fetch_users(self.config_model.wiki_config, ExportModel, ExportController)

    def create_users_workbook(self):
        user_model = self.user_model
        users = self.users
        workbook_controller = WorkbookController()
        with closing(workbook_controller.create_users_workbook(
            user_model.field_title, user_model.fields, user_model.titles, users
        )) as workbook:
            workbook_buffer = BytesIO()
            workbook.save(workbook_buffer)
            workbook_buffer = workbook_controller.fix_workbook_mime_type(workbook_buffer)
            self.workbook_buffer = workbook_buffer

    def write_users_workbook(self):
        with open(self.config_model.users_excel_file_path, "wb") as file:
            file.write(self.workbook_buffer.getvalue())

    def upload_users_workbook(self):
        # file_data = Path(self.users_excel_file_path).read_bytes()
        file_data = self.workbook_buffer.getvalue()
        wiki_controller = self.wiki_controller
        result = wiki_controller.upload_file(self.config_model.users_excel_file_name, file_data)
        uri = result["upload"]["imageinfo"]["descriptionurl"]
        print(uri)


def main(*args):
    argument_controller = ArgumentController(CONFIG)
    config = argument_controller.create_config_from_system_arguments()
    config_model = ConfigModel(config)

    main_controller = MainController(config_model)
    main_controller.run()


if __name__ == "__main__":
    main()